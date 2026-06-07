import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Define data directory
TASK_DIR = '/home/deepakraj29/Desktop/intern/Task 3'
if not os.path.exists(TASK_DIR):
    os.makedirs(TASK_DIR)

OUTPUT_FILE = os.path.join(TASK_DIR, 'Volunteer_Opportunities_India.xlsx')

# 1. Prepare dataset of 16 opportunities
data = [
    {
        "Opportunity Name": "Teach For India Fellowship",
        "Organization": "Teach For India",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Graduates or young professionals of any discipline. Must be Indian citizens or OCI.",
        "Duration & Stipend": "2 Years | ₹23,000/month + housing allowance",
        "Application Link": "https://www.teachforindia.org/",
        "Deadline / Cycle": "Annual cohort (applications open around Aug-Sep for next year's cycle)",
        "Brief Description": "A two-year full-time leadership development program where fellows serve as teachers in under-resourced schools to drive educational equity."
    },
    {
        "Opportunity Name": "SBI Youth for India Fellowship",
        "Organization": "SBI Foundation",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Bachelor's degree completed by program start; aged 21-32. Citizens of India, Nepal, Bhutan, or OCI.",
        "Duration & Stipend": "13 Months | ₹16,000/month stipend + travel & readjustment grants",
        "Application Link": "https://www.youthforindia.org/",
        "Deadline / Cycle": "Annual cohort (applications open around Feb-May)",
        "Brief Description": "A rural development fellowship enabling youth to work on grassroots projects across 12 thematic areas with partner NGOs."
    },
    {
        "Opportunity Name": "Gandhi Fellowship",
        "Organization": "Piramal Foundation",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Final year UG/PG students or graduates aged below 26, showing leadership traits.",
        "Duration & Stipend": "23 Months | ₹14,000/month stipend + boarding/lodging support",
        "Application Link": "https://gandhifellowship.org/",
        "Deadline / Cycle": "Annual cohort (admissions close around March-April)",
        "Brief Description": "A residential leadership journey placing fellows in public schools to drive system change and develop personal emotional intelligence."
    },
    {
        "Opportunity Name": "Bhumi Fellowship",
        "Organization": "Bhumi (NGO)",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Graduates or young professionals aged 20-30. Strong leadership potential and empathy.",
        "Duration & Stipend": "2 Years | ₹18,000/month grant",
        "Application Link": "https://fellowship.bhumi.ngo/",
        "Deadline / Cycle": "Annual cohort (applications close around April-May)",
        "Brief Description": "A leadership development program enabling youth to transform government primary school ecosystems and improve learning outcomes."
    },
    {
        "Opportunity Name": "Prime Minister's Research Fellowship (PMRF)",
        "Organization": "Ministry of Education, GoI",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Grads from IITs, IISc, IISERs, NITs with CGPA >= 8.0, or selected in PhD at hosting institutions.",
        "Duration & Stipend": "Up to 5 Years | ₹70,000 - ₹80,000/month + ₹2 Lakhs annual research grant",
        "Application Link": "https://pmrf.gov.in/",
        "Deadline / Cycle": "Biannual cycles (typically May and December)",
        "Brief Description": "Flagship doctoral scheme attracting engineering and science scholars to carry out research at premier Indian institutions."
    },
    {
        "Opportunity Name": "JNCASR Summer Research Fellowship (SRFP)",
        "Organization": "Jawaharlal Nehru Centre for Advanced Scientific Research",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Undergraduate (UG) and Postgraduate (PG) science/engineering students.",
        "Duration & Stipend": "2 Months | ₹10,000/month stipend + travel allowance",
        "Application Link": "http://www.jncasr.ac.in/fe/srfp.php",
        "Deadline / Cycle": "Annual (opens in Nov, closes in Dec/Jan)",
        "Brief Description": "Prestigious summer research program matching students with leading faculty in biotechnology, materials science, and physics."
    },
    {
        "Opportunity Name": "Science Academies' Summer Research Fellowship",
        "Organization": "Indian Academy of Sciences (IAS)",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Science and engineering students/teachers enrolled in Indian colleges/universities.",
        "Duration & Stipend": "8 Weeks | Monthly stipend + round-trip travel allowance",
        "Application Link": "https://web-japps.ias.ac.in/",
        "Deadline / Cycle": "Annual (applications close in November)",
        "Brief Description": "Summer research training matching candidates with host scientists belonging to the three national science academies."
    },
    {
        "Opportunity Name": "NITI Aayog Internship Scheme",
        "Organization": "NITI Aayog",
        "Opportunity Type": "Internship",
        "Eligibility": "UG/PG students or research scholars from recognized universities in India/abroad. High academic marks (e.g. 85% in 12th).",
        "Duration & Stipend": "1 to 3 Months | Unpaid",
        "Application Link": "https://niti.gov.in/internship",
        "Deadline / Cycle": "Open from the 1st to the 10th of every month",
        "Brief Description": "Internship at the government's premier policy think tank, providing insights into policy analysis and governance."
    },
    {
        "Opportunity Name": "Ministry of External Affairs (MEA) Internship",
        "Organization": "Ministry of External Affairs, GoI",
        "Opportunity Type": "Internship",
        "Eligibility": "Indian citizens pursuing/completed graduation; maximum age of 25 years.",
        "Duration & Stipend": "1 to 3 Months | ₹10,000/month stipend + flight tickets support",
        "Application Link": "https://www.internship.mea.gov.in/",
        "Deadline / Cycle": "Twice a year (cycles: Apr-Sep and Oct-Mar)",
        "Brief Description": "Provides first-hand exposure to India's foreign policy formulation, bilateral relations, and multilateral diplomacy."
    },
    {
        "Opportunity Name": "Make A Difference (MAD) Volunteering",
        "Organization": "Make A Difference",
        "Opportunity Type": "Volunteer",
        "Eligibility": "High school/college students or youth. Selection based on interview and training.",
        "Duration & Stipend": "1 Year (weekend commitment, ~2-4 hours/week) | Unpaid",
        "Application Link": "https://makeadiff.in/join/",
        "Deadline / Cycle": "Rolling applications (hiring peaks in June-August)",
        "Brief Description": "Volunteers mentor and teach children in shelter homes, building academic foundations and lifeskills support."
    },
    {
        "Opportunity Name": "Bhumi Volunteer Squad",
        "Organization": "Bhumi (NGO)",
        "Opportunity Type": "Volunteer",
        "Eligibility": "Open to all students and professionals. No age or background barrier.",
        "Duration & Stipend": "Flexible (weekend activities) | Unpaid",
        "Application Link": "https://bhumi.ngo/volunteer/",
        "Deadline / Cycle": "Open year-round",
        "Brief Description": "Offers flexible weekend volunteering opportunities in beach cleanups, educational programs, and blood donation drives."
    },
    {
        "Opportunity Name": "CRY Volunteering & Internships",
        "Organization": "Child Rights and You (CRY)",
        "Opportunity Type": "Volunteer / Internship",
        "Eligibility": "UG/PG students or working professionals interested in child advocacy.",
        "Duration & Stipend": "Flexible (Volunteering) / 4-8 Weeks (Internship) | Unpaid",
        "Application Link": "https://www.cry.org/volunteer/",
        "Deadline / Cycle": "Open year-round",
        "Brief Description": "Engagement campaigns focusing on child education, child labour prevention, health drives, and advocacy."
    },
    {
        "Opportunity Name": "Goonj Internship & Volunteering",
        "Organization": "Goonj (NGO)",
        "Opportunity Type": "Volunteer / Internship",
        "Eligibility": "College students and general public interested in community development.",
        "Duration & Stipend": "1 Month (150 hours commitment for internships) | Unpaid",
        "Application Link": "https://goonj.org/internship/",
        "Deadline / Cycle": "Rolling applications",
        "Brief Description": "Internship focusing on urban-rural circular economy, recycling logistics, disaster relief management, and campaigns."
    },
    {
        "Opportunity Name": "LAMP Fellowship",
        "Organization": "PRS Legislative Research",
        "Opportunity Type": "Fellowship",
        "Eligibility": "Indian citizens aged 25 or below with at least a Bachelor's degree in any stream.",
        "Duration & Stipend": "11 Months | ₹20,000/month stipend",
        "Application Link": "https://www.prsindia.org/lamp",
        "Deadline / Cycle": "Annual cohort (applications open around October-December)",
        "Brief Description": "Legislative Assistants to Members of Parliament (LAMP) Fellowship, pairing fellows with MPs to assist in parliamentary research."
    },
    {
        "Opportunity Name": "Reliance Foundation UG Scholarship",
        "Organization": "Reliance Foundation",
        "Opportunity Type": "Scholarship",
        "Eligibility": "First-year undergraduate students in any stream with household income < ₹15 Lakhs/year.",
        "Duration & Stipend": "Full course duration | Up to ₹2 Lakhs total financial grant",
        "Application Link": "https://www.reliancefoundation.org/",
        "Deadline / Cycle": "Annual selection (applications open in September-November)",
        "Brief Description": "Merit-cum-means scholarship supporting meritorious students to cover college tuition, books, and living expenses."
    },
    {
        "Opportunity Name": "Reliance Foundation PG Scholarship",
        "Organization": "Reliance Foundation",
        "Opportunity Type": "Scholarship",
        "Eligibility": "First-year PG students in future-ready tech streams (AI, CS, Life Sciences, etc.).",
        "Duration & Stipend": "Full course duration | Up to ₹6 Lakhs total scholarship grant",
        "Application Link": "https://www.reliancefoundation.org/",
        "Deadline / Cycle": "Annual selection (applications open in September-November)",
        "Brief Description": "Prestigious scholarship supporting postgraduate students conducting research and development in cutting-edge tech."
    }
]

df = pd.DataFrame(data)

# 2. Compile using openpyxl for premium formatting
wb = Workbook()
ws = wb.active
ws.title = "Volunteer Opportunities India"

# Show gridlines
ws.views.sheetView[0].showGridLines = True

# Convert data to workbook rows
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Define design styles
navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
cell_font = Font(name="Arial", size=10, color="1F2937")

thin_border_side = Side(border_style="thin", color="E5E7EB")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

left_align = Alignment(horizontal="left", vertical="center")
center_align = Alignment(horizontal="center", vertical="center")
wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Define column widths manually for best visual presentation
col_widths = {
    "A": 28,  # Opportunity Name
    "B": 28,  # Organization
    "C": 18,  # Opportunity Type
    "D": 32,  # Eligibility (wrapped)
    "E": 26,  # Duration & Stipend
    "F": 30,  # Application Link
    "G": 24,  # Deadline / Cycle
    "H": 40   # Brief Description (wrapped)
}

# Style Header Row (Row 1)
for col_num in range(1, 9):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = navy_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[1].height = 28

# Style Data Rows
for row_num in range(2, len(data) + 2):
    is_odd = (row_num % 2 == 1)
    ws.row_dimensions[row_num].height = 50  # Give tall height to accommodate wrapped text
    
    for col_num in range(1, 9):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = cell_font
        cell.border = thin_border
        
        # Apply alternating background fills (zebra striping)
        if is_odd:
            cell.fill = zebra_fill
            
        # Alignments
        if col_num in [1, 2, 5]:  # Name, Organization, Duration
            cell.alignment = left_align
        elif col_num in [3, 6, 7]:  # Type, Link, Deadline
            cell.alignment = center_align
        elif col_num in [4, 8]:  # Eligibility, Description (Wrapped)
            cell.alignment = wrap_left_align

# Apply fixed column dimensions
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save the workbook
wb.save(OUTPUT_FILE)
print(f"Excel file successfully created and styled at: {OUTPUT_FILE}")
